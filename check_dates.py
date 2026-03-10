#!/usr/bin/env python3
import os
from openpyxl import load_workbook
from datetime import datetime, timedelta
import glob

# Directory containing the Excel files
directory = "창고별재고-january"

# Get all Excel files
excel_files = sorted(glob.glob(os.path.join(directory, "*.xlsx")))

print(f"Found {len(excel_files)} Excel files\n")

# Dictionary to store filename -> date mapping
file_dates = {}

# Read first row of each file
for filepath in excel_files:
    filename = os.path.basename(filepath)
    try:
        # Load workbook and get first sheet
        wb = load_workbook(filepath, read_only=True, data_only=True)
        ws = wb.active

        # Get first row
        first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        date_found = None

        for value in first_row:
            if value is not None:
                # Try to parse as date
                if isinstance(value, datetime):
                    date_found = value
                    break
                elif isinstance(value, str):
                    # Try to parse string as date
                    for fmt in ["%Y-%m-%d", "%Y/%m/%d", "%m/%d/%Y", "%d/%m/%Y", "%Y%m%d", "%Y.%m.%d"]:
                        try:
                            date_found = datetime.strptime(value, fmt)
                            break
                        except:
                            continue
                    if date_found:
                        break

        wb.close()

        if date_found:
            file_dates[filename] = date_found
            print(f"{filename}: {date_found.strftime('%Y-%m-%d')}")
        else:
            print(f"{filename}: NO DATE FOUND - First row: {first_row}")

    except Exception as e:
        print(f"{filename}: ERROR - {str(e)}")

print("\n" + "="*80)
print("ANALYSIS:")
print("="*80)

# Check for missing dates in January 2025
if file_dates:
    dates = sorted(file_dates.values())
    print(f"\nDate range: {dates[0].strftime('%Y-%m-%d')} to {dates[-1].strftime('%Y-%m-%d')}")

    # Determine which year and month to check
    year = dates[0].year
    month = dates[0].month

    # Generate all dates in January
    start_date = datetime(year, month, 1)
    if month == 12:
        end_date = datetime(year + 1, 1, 1)
    else:
        end_date = datetime(year, month + 1, 1)

    all_january_dates = []
    current = start_date
    while current < end_date:
        all_january_dates.append(current)
        current += timedelta(days=1)

    # Check which dates are missing
    found_dates = set(d.date() for d in file_dates.values())
    missing_dates = [d for d in all_january_dates if d.date() not in found_dates]

    if missing_dates:
        print(f"\nMISSING DATES ({len(missing_dates)}):")
        for date in missing_dates:
            print(f"  - {date.strftime('%Y-%m-%d')}")
    else:
        print(f"\nAll {len(all_january_dates)} dates in {start_date.strftime('%B %Y')} are present!")

print("\n" + "="*80)
print("FILE -> DATE MAPPING:")
print("="*80)
for filename, date in sorted(file_dates.items(), key=lambda x: x[1]):
    print(f"{filename} -> {date.strftime('%Y-%m-%d')}")
